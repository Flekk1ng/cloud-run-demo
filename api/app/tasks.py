import os
import pandas as pd
from openpyxl.styles import numbers

from app import constants as c
from app import utils
from app import xls_styles


def run_tasks(*, report_type_country_filenames, output_filename):
    task_one(report_type=c.REPORT_TYPES[0], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[0]], output_filename=output_filename)
    task_two(report_type=c.REPORT_TYPES[1], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[1]], output_filename=output_filename)
    task_three(report_type=c.REPORT_TYPES[2], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[2]], output_filename=output_filename)
    task_four(report_type=c.REPORT_TYPES[3], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[3]], output_filename=output_filename)
    task_five(report_type=c.REPORT_TYPES[4], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[4]], output_filename=output_filename)
    task_six(report_type=c.REPORT_TYPES[5], country_file_names_for_task=report_type_country_filenames[c.REPORT_TYPES[5]], output_filename=output_filename)

def task_one(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование первой страницы отчета (для country_monthly_data) 
    """
    dataframes = {}
    
    for country, filename in country_file_names_for_task.items():
        df = pd.read_csv(os.path.join(utils.get_upload_path(), filename), header=None)
        dataframes[country] = df.T

    writer = get_writer(output_filename=output_filename, new=True)

    for i in range(1, 6):
        df = pd.DataFrame()
        first_country = list(dataframes.keys())[0]
        df.insert(0, 0, dataframes[first_country][0])
        df[0][0] = dataframes[first_country][i][0]

        pos = 1
        for country in dataframes.keys():
            df.insert(pos, pos, dataframes[country][i])
            df[pos][0] = country
            pos += 1

        df.T.to_excel(writer, sheet_name=report_type, startrow=(i - 1) * (len(dataframes.keys()) + 3), index=False, header=False)

    xls_styles.set_styles(writer, report_type, [0])

    writer.close()


def task_two(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование второй страницы отчета (для country_yearly_data) 
    """
    dataframes = {}

    for country, filename in country_file_names_for_task.items():
        df = pd.read_csv(os.path.join(utils.get_upload_path(), filename), header=None)
        dataframes[country] = df.T
    
    writer = get_writer(output_filename=output_filename, new=False)

    for i in range(1, 7):
        df = pd.DataFrame()
        first_country = list(dataframes.keys())[0]
        df.insert(0, 0, pd.concat([dataframes[first_country][0], pd.Series([''])], ignore_index=True))
        df[0][0] = dataframes[first_country][i][0]
        df[0][6] = 'Same period of the Year before'
        df[0][7] = 'Current year available period'
        df[0][8] = 'Current year available period'
        
        pos = 1
        for country in dataframes.keys():
            series = dataframes[country][i]
            first_value = series[0]
            if 'rate' in dataframes[first_country][i][0]:
                other_values = series[1:].astype(float) / 100
            else:
                other_values = series[1:].astype(float)
            updated_series = pd.Series([first_value] + other_values.tolist())
            df.insert(pos, pos, updated_series)

            df[pos][0] = country
            df[pos][8] = dataframes[country][0][7]
            pos += 1

        start = (i - 1) * (len(dataframes.keys()) + 3)
        df.T.to_excel(writer, sheet_name=report_type, startrow=start, index=False, header=False)

        ws = writer.sheets[report_type]

        for col_num in range(2, ws.max_row + 1):
            if 'rate' in dataframes[first_country][i][0]:
                for row_num in range(start + 2, start + (len(dataframes.keys()) + 2)):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

    xls_styles.set_styles(writer, report_type, list(range(9)))

    writer.close()


def task_three(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование третьей страницы отчета (для global_data_product) 
    """
    
    file_names_for_task = [file_name_for_task for _, file_name_for_task in country_file_names_for_task.items()]
    
    dataframe = pd.read_csv(os.path.join(utils.get_upload_path(), file_names_for_task[0]), header=None).T

    for i in range(1, 5):
        if 'rate' in dataframe[i][0]:
            series = dataframe[i]
            first_value = series[0]
            other_values = series[1:].astype(float) / 100
            updated_series = pd.Series([first_value] + other_values.tolist())
            dataframe[i] = updated_series
    
    writer = get_writer(output_filename=output_filename, new=False)
    
    dataframe.T.to_excel(writer, sheet_name=report_type, index=False, header=False)
    
    ws = writer.sheets[report_type]

    for col_num in range(2, ws.max_row + 2):
        for row_num in range(4, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    xls_styles.set_styles(writer, report_type, [0])

    writer.close()


def task_four(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование четвертой страницы отчета (для indicators_country) 
    """
    dataframes = {}

    for country, filename in country_file_names_for_task.items():
        df = pd.read_csv(os.path.join(utils.get_upload_path(), filename), header=None)
        dataframes[country] = df

    writer = get_writer(output_filename=output_filename, new=False)

    df = pd.DataFrame()
    first_country = list(dataframes.keys())[0]
    df.insert(0, 0, pd.concat([pd.Series(['']), dataframes[first_country][0]], ignore_index=True))

    perc_rows = [3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 18, 19, 21, 22, 23, 24, 26, 27, 30, 31]
    perc_rows = [x + 1 for x in perc_rows]

    pos = 1
    for country in dataframes.keys():
        series = pd.concat([pd.Series([country]), dataframes[country][1]], ignore_index=True)
        new_series = series.copy()
        new_series[3:-5] = new_series[3:-5].astype(float)
        new_series.iloc[perc_rows] = new_series.iloc[perc_rows] / 100
        df.insert(pos, pos, new_series)
        pos += 1

    df.to_excel(writer, sheet_name=report_type, index=False, header=False)

    ws = writer.sheets[report_type]

    for col_num in range(2, ws.max_row + 2):
        for row_num in [r + 1 for r in perc_rows]:
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    xls_styles.set_styles(writer, report_type, list(range(len(dataframes.keys()) + 1)))

    writer.close()


def task_five(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование пятой страницы отчета (для suppliers_full_year) 
    """
    dataframe = pd.DataFrame()
    for country, filename in country_file_names_for_task.items():
        df = pd.read_csv(os.path.join(utils.get_upload_path(), filename))

        reporter_series = pd.Series([country] * len(df), name='ReporterCode')

        # Вставляем Series после первого столбца
        df.insert(1, 'ReporterCode', reporter_series)

        dataframe = pd.concat([dataframe, df])
    
    writer = get_writer(output_filename=output_filename, new=False)

    suppliers_column_name = dataframe.columns[0]
    unique_suppliers = dataframe[suppliers_column_name].unique()

    new_final_dataframe = pd.DataFrame()
    for supplier in unique_suppliers:
        group_df = dataframe[dataframe[suppliers_column_name] == supplier]
        
        del group_df[suppliers_column_name]

        group_df = group_df.reset_index(drop=True)

        # Создание нового DataFrame с первой ячейкой, содержащей значение группировки
        new_df = pd.DataFrame({suppliers_column_name: [supplier] + [''] * (len(group_df.index) - 1)})
        
        # Добавляем все серии из group_df к new_df
        new_df = new_df.assign(**group_df)

        total_row = new_df.sum()
        total_row.iloc[0] = ''
        total_df = pd.DataFrame(total_row)

        total_row['ReporterCode'] = 'Total for ' + str(supplier)
        new_df = pd.concat([new_df, total_df.T], ignore_index=True)

        new_df['Share in total imports of $, %'] = new_df['Share in total imports of $, %'] / 100
        new_df['Share in total imports of tons, %'] = new_df['Share in total imports of tons, %'] / 100
        
        new_final_dataframe = pd.concat([new_final_dataframe, new_df])

    new_final_dataframe.to_excel(writer, sheet_name=report_type, index=False, header=True)

    xls_styles.set_styles(writer, report_type, list(range(0, 7)))

    ws = writer.sheets[report_type]

    for row_num in range(1, ws.max_row + 1):
        for col_num in range(6, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    for row_num in range(1, len(new_final_dataframe.index) + 1):
        if 'Total for' in new_final_dataframe.iloc[row_num - 1, 1]:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num + 1, column=col)
                cell.font = cell.font.copy(bold=True)

    writer.close()


def task_six(*, report_type, country_file_names_for_task, output_filename):
    """
    Формирование шестой страницы отчета (для suppliers_ltm) 
    """
    dataframe = pd.DataFrame()

    suppliers_column_name = 'Suppliers in LTM'

    for country, filename in country_file_names_for_task.items():
        df = pd.read_csv(os.path.join(utils.get_upload_path(), filename))

        period = df.columns[0]

        start_index = period.index('Suppliers in LTM ')

        end_index = start_index + len('Suppliers in LTM ')

        period = period[:start_index] + period[end_index:]

        suppliers_in_ltm_period_series = pd.Series([period] * len(df))
        df['Suppliers in LTM Period'] = suppliers_in_ltm_period_series
        df = df.rename(columns={df.columns[0]: suppliers_column_name})

        reporter_series = pd.Series([country] * len(df), name='ReporterCode')

        # Вставляем Series после первого столбца
        df.insert(1, 'ReporterCode', reporter_series)

        dataframe = pd.concat([dataframe, df])
    
    writer = get_writer(output_filename=output_filename, new=False)

    unique_suppliers = dataframe[suppliers_column_name].unique()

    new_final_dataframe = pd.DataFrame()
    for supplier in unique_suppliers:
        group_df = dataframe[dataframe[suppliers_column_name] == supplier]
        
        del group_df[suppliers_column_name]

        group_df = group_df.reset_index(drop=True)

        # Создание нового DataFrame с первой ячейкой, содержащей значение группировки
        new_df = pd.DataFrame({suppliers_column_name: [supplier] + [''] * (len(group_df.index) - 1)})
        
        # Добавляем все серии из group_df к new_df
        new_df = new_df.assign(**group_df)

        total_row = new_df.sum()
        total_row.iloc[0] = ''
        total_row.iloc[-1] = ''
        total_df = pd.DataFrame(total_row)

        total_row['ReporterCode'] = 'Total for ' + str(supplier)
        new_df = pd.concat([new_df, total_df.T], ignore_index=True)

        new_df['share of total imports in US $'] = new_df['share of total imports in US $'] / 100
        new_df['share of total imports in tons'] = new_df['share of total imports in tons'] / 100
        
        new_final_dataframe = pd.concat([new_final_dataframe, new_df])

    new_final_dataframe.to_excel(writer, sheet_name=report_type, index=False, header=True)

    xls_styles.set_styles(writer, report_type, list(range(0, 10)))

    ws = writer.sheets[report_type]

    for row_num in range(1, ws.max_row + 1):
        for col_num in range(8, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    for row_num in range(1, len(new_final_dataframe.index) + 1):
        if 'Total for' in new_final_dataframe.iloc[row_num - 1, 1]:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num + 1, column=col)
                cell.font = cell.font.copy(bold=True)

    writer.close()


def get_writer(*, output_filename, new: bool) -> pd.ExcelWriter:
    download_file_path = os.path.join(utils.get_download_path(), output_filename)

    if (new):
        writer = pd.ExcelWriter(f"{str(download_file_path)}.xlsx", engine='openpyxl', mode='w')
    else:
        writer = pd.ExcelWriter(f"{str(download_file_path)}.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay')

    return writer