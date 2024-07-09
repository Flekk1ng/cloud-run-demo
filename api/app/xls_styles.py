from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def set_styles(writer, sheet_name: str, required_columns: list):
    ws = writer.sheets[sheet_name]

    font = Font(name='Arial', size=10, bold=False)
    alignment = Alignment(horizontal='right', vertical='bottom')

    fields_alignment = Alignment(horizontal='left', vertical='bottom')

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = font
        cell.alignment = fields_alignment

    for row in range(1, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = alignment

    all_columns = list(ws.columns)

    for col_idx in required_columns:
        column = all_columns[col_idx]
        column_letter = get_column_letter(col_idx + 1)
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) * 0.9 > max_length:
                    max_length = len(str(cell.value)) * 0.9
            except:
                pass
        adjusted_width = max_length
        ws.column_dimensions[column_letter].width = adjusted_width
